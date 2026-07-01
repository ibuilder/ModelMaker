"""Generic Excel/CSV import for ANY config-driven module — the biggest data-entry / adoption lever.
Two-step: **preview** (parse the sheet, auto-suggest a column->field mapping, coerce a sample, flag
unmapped required fields) then **import** (validate required, type-coerce, batch-create). Power users
seed records from a spreadsheet instead of the one-by-one form. Reuses openpyxl (already vendored);
CSV via stdlib. Portable across SQLite/Postgres (writes through the module engine)."""
from __future__ import annotations

import csv
import io
import re
from datetime import datetime
from typing import Any

_MAX_ROWS = 10_000                                   # cap created records per import (DoS guard)
_MAX_SCAN = 400_000                                  # hard cap on rows scanned
# field types we can import into (rollup/signature are computed/interactive -> skipped)
IMPORTABLE_TYPES = {"text", "number", "currency", "date", "textarea", "select", "multiselect", "reference"}


def _norm(s: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s or "").strip().lower())


def importable_fields(key: str) -> list[dict[str, Any]]:
    from . import modules as me
    mod = me.get_module(key)
    out = []
    for f in mod.get("fields", []):
        if f.get("type", "text") in IMPORTABLE_TYPES:
            out.append({"name": f["name"], "label": f.get("label", f["name"]),
                        "type": f.get("type", "text"), "required": bool(f.get("required")),
                        "options": f.get("options")})
    return out


def parse_table(data: bytes, filename: str | None) -> tuple[list[str], list[list]]:
    """Return (headers, body_rows). First non-empty row is the header. Handles .xlsx and .csv."""
    name = (filename or "").lower()
    rows: list[list] = []
    if name.endswith((".xlsx", ".xlsm", ".xls")):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        for r in ws.iter_rows(values_only=True):
            rows.append(["" if c is None else c for c in r])
            if len(rows) >= _MAX_SCAN:
                break
        wb.close()
    else:                                            # default to CSV (also handles .txt/tsv-ish)
        text = data.decode("utf-8-sig", errors="replace")
        for r in csv.reader(io.StringIO(text)):
            rows.append(r)
            if len(rows) >= _MAX_SCAN:
                break
    hi = next((i for i, r in enumerate(rows) if any(str(c).strip() for c in r)), None)
    if hi is None:
        return [], []
    headers = [str(c).strip() for c in rows[hi]]
    body = [r for r in rows[hi + 1:] if any(str(c).strip() for c in r)]
    return headers, body


def suggest_mapping(headers: list[str], fields: list[dict]) -> dict[str, str]:
    """Auto-map each source column to a field by exact name/label match (normalized)."""
    by_norm: dict[str, str] = {}
    for f in fields:                                 # label first so name wins ties
        by_norm.setdefault(_norm(f["label"]), f["name"])
        by_norm[_norm(f["name"])] = f["name"]
    out: dict[str, str] = {}
    used: set[str] = set()
    for h in headers:
        fld = by_norm.get(_norm(h))
        if fld and fld not in used:
            out[h] = fld
            used.add(fld)
    return out


def _coerce(val: Any, ftype: str) -> Any:
    s = "" if val is None else str(val).strip()
    if s == "":
        return None
    if ftype in ("number", "currency"):
        try:
            return float(re.sub(r"[,$%\s]", "", s))
        except ValueError:
            return None
    if ftype == "date":
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s[:10], fmt).date().isoformat()
            except ValueError:
                pass
        return s                                     # leave as-is if unrecognised
    if ftype == "multiselect":
        return [p.strip() for p in re.split(r"[;,|]", s) if p.strip()]
    return s


def _row_to_data(headers: list[str], row: list, mapping: dict[str, str], ftype: dict[str, str]) -> dict:
    rec: dict[str, Any] = {}
    for i, h in enumerate(headers):
        fld = mapping.get(h)
        if fld and i < len(row):
            v = _coerce(row[i], ftype.get(fld, "text"))
            if v is not None and v != []:
                rec[fld] = v
    return rec


def preview(key: str, data: bytes, filename: str | None, mapping: dict | None = None) -> dict[str, Any]:
    fields = importable_fields(key)
    headers, body = parse_table(data, filename)
    sugg = mapping or suggest_mapping(headers, fields)
    ftype = {f["name"]: f["type"] for f in fields}
    sample = [_row_to_data(headers, row, sugg, ftype) for row in body[:8]]
    req = [f["name"] for f in fields if f["required"]]
    mapped = set(sugg.values())
    return {"headers": headers, "fields": fields, "suggested_mapping": sugg,
            "row_count": len(body), "sample": sample,
            "unmapped_required": [r for r in req if r not in mapped]}


def do_import(db, key: str, pid: str, data: bytes, filename: str | None,
              mapping: dict[str, str], actor: str, party: str | None) -> dict[str, Any]:
    from . import modules as me
    fields = importable_fields(key)
    headers, body = parse_table(data, filename)
    ftype = {f["name"]: f["type"] for f in fields}
    req = [f["name"] for f in fields if f["required"]]
    created = 0
    errors: list[dict] = []
    for n, row in enumerate(body[:_MAX_ROWS], start=1):
        rec = _row_to_data(headers, row, mapping, ftype)
        if not rec:
            continue                                 # wholly blank / unmapped row
        missing = [r for r in req if not rec.get(r)]
        if missing:
            errors.append({"row": n, "error": f"missing required: {', '.join(missing)}"})
            continue
        try:
            me.create_record(db, key, pid, {"data": rec}, actor, party)
            created += 1
        except Exception as e:                       # never let one bad row abort the batch
            errors.append({"row": n, "error": str(e)[:160]})
    return {"imported": created, "error_count": len(errors), "errors": errors[:50],
            "truncated": len(body) > _MAX_ROWS}


def template_csv(key: str) -> str:
    """A header-only CSV (the module's importable field labels) for users to fill in and re-upload."""
    fields = importable_fields(key)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([f["label"] + (" *" if f["required"] else "") for f in fields])
    return buf.getvalue()
